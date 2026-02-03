from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
from collections import defaultdict
from copy import deepcopy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


ALLOWED_SECTIONS = {"MIE", "EEE", "CAE"}
ALLOWED_NATURE = {"Theory", "Theory with Lab", "Practical"}
ALLOWED_LEVEL = {"Bachelor", "Advanced Diploma", "Diploma Second Year", "Diploma First Year"}
ALLOWED_INCLUDED = {"MID", "FINAL", "MID & Final"}


@dataclass
class CLOEntry:
    clo_no: int
    clo_text: str
    contact_hours: float = 0.0
    included_for: str = ""  # "MID" | "FINAL" | "MID & Final" | ""


@dataclass
class CourseOutcomesModel:
    semester: str = ""          # "1" | "2" | "3"
    ay: str = ""
    section: str = ""           # "MIE" | "EEE" | "CAE"
    course_code: str = ""
    course_title: str = ""
    nature_of_course: str = ""  # Theory / Theory with Lab / Practical
    level: str = ""             # Bachelor / ...
    mid_exam_marks: Optional[float] = None
    final_exam_marks: Optional[float] = None
    template_used_for: str = "" # "MID" | "FINAL"
    clos: List[CLOEntry] = field(default_factory=list)


# ----------------------------
# Extraction helpers
# ----------------------------

def _semester_to_number(sem: str) -> str:
    s = (sem or "").strip().lower()
    if "i" in s and "semester" in s:
        # "Semester I"
        if "iii" in s:
            return "3"
        if "ii" in s:
            return "2"
        return "1"
    # already "1"/"2"/"3"
    if s in {"1", "2", "3"}:
        return s
    return ""


def _infer_nature_of_course(cdp: Dict[str, Any]) -> str:
    practical = cdp.get("practical_df") or []
    has_practical_hours = any(float(r.get("hours", 0) or 0) > 0 for r in practical if isinstance(r, dict))
    return "Theory with Lab" if has_practical_hours else "Theory"


def _map_section(doc_section: str, mapping: Dict[str, str] | None = None) -> str:
    if not doc_section:
        return ""
    if mapping and doc_section in mapping:
        return mapping[doc_section]
    # safe default: only accept if already allowed
    return doc_section if doc_section in ALLOWED_SECTIONS else ""


def _compute_contact_hours_per_clo(
    theory_df: List[Dict[str, Any]],
    practical_df: List[Dict[str, Any]],
    split_across_clos: bool = True
) -> Dict[str, float]:
    """
    Returns dict like {"CLO1": 10.5, "CLO2": 8.0, ...}
    Uses hours from theory_df/practical_df rows and their 'clos' list.
    If split_across_clos=True: divide row hours equally among all CLOs on that row.
    """
    out = defaultdict(float)

    def consume(rows: List[Dict[str, Any]]):
        for r in rows:
            if not isinstance(r, dict):
                continue
            try:
                h = float(r.get("hours", 0) or 0)
            except Exception:
                continue
            if h <= 0:
                continue
            clos = r.get("clos") or []
            if isinstance(clos, str):
                clos = [clos]
            clos = [c.strip() for c in clos if isinstance(c, str) and c.strip()]
            if not clos:
                continue

            if split_across_clos:
                per = h / len(clos)
                for c in clos:
                    out[c] += per
            else:
                for c in clos:
                    out[c] += h

    consume(theory_df or [])
    consume(practical_df or [])
    return dict(out)


def extract_course_outcomes_from_cdp(
    cdp_json: Dict[str, Any],
    *,
    default_level: str = "Bachelor",
    section_mapping: Dict[str, str] | None = None,
    split_contact_hours: bool = True
) -> CourseOutcomesModel:
    """
    Extracts the staff-owned COURSE_OUTCOMES sheet inputs from an approved CDP JSON.
    """
    doc = cdp_json.get("doc", {}) or {}
    course = cdp_json.get("course", {}) or {}

    semester = _semester_to_number(str(doc.get("semester", "")))
    ay = str(doc.get("academic_year", "") or "")
    section = _map_section(str(doc.get("section", "") or ""), mapping=section_mapping)

    course_code = str(course.get("course_code", "") or "")
    course_title = str(course.get("course_title", "") or "")

    nature = _infer_nature_of_course(cdp_json)
    if nature not in ALLOWED_NATURE:
        nature = "Theory with Lab"

    level = default_level if default_level in ALLOWED_LEVEL else ""

    # CLO texts
    clos_df = cdp_json.get("clos_df") or []
    clo_texts = []
    for row in clos_df:
        if isinstance(row, dict):
            t = str(row.get("learning_outcomes", "") or "").strip()
            if t:
                clo_texts.append(t)

    # contact hours per CLO from weekly distributions
    theory_df = cdp_json.get("theory_df") or []
    practical_df = cdp_json.get("practical_df") or []
    clo_hours = _compute_contact_hours_per_clo(theory_df, practical_df, split_across_clos=split_contact_hours)

    clos: List[CLOEntry] = []
    for i, text in enumerate(clo_texts, start=1):
        key = f"CLO{i}"
        hours = float(clo_hours.get(key, 0.0) or 0.0)
        clos.append(CLOEntry(clo_no=i, clo_text=text, contact_hours=round(hours, 2), included_for=""))

    return CourseOutcomesModel(
        semester=semester,
        ay=ay,
        section=section,
        course_code=course_code,
        course_title=course_title,
        nature_of_course=nature,
        level=level,
        mid_exam_marks=None,
        final_exam_marks=None,
        template_used_for="",  # staff selects MID/FINAL in UI
        clos=clos
    )


# ----------------------------
# Computation (UI preview)
# ----------------------------

def _norm_included(val: str) -> str:
    v = (val or "").strip()
    # normalize common variants
    v = v.replace("&", "&").replace("and", "&")
    if v.lower() in {"mid", "midsem", "mid sem"}:
        return "MID"
    if v.lower() in {"final"}:
        return "FINAL"
    if v.lower() in {"mid & final", "mid&final", "mid and final"}:
        return "MID & Final"
    return v if v in ALLOWED_INCLUDED else ""


def compute_tos_numbers(model: CourseOutcomesModel) -> Dict[str, Any]:
    """
    Computes the yellow/derived columns for UI display (Excel will compute via formulas).
    Rounding: weight % => 0 decimals; weighted marks => 2 decimals.
    """
    mid_marks = model.mid_exam_marks
    final_marks = model.final_exam_marks

    rows = []
    for clo in model.clos:
        inc = _norm_included(clo.included_for)
        rows.append({
            "clo_no": clo.clo_no,
            "clo_text": clo.clo_text,
            "contact_hours": float(clo.contact_hours or 0),
            "included_for": inc,
        })

    mid_set = [r for r in rows if r["included_for"] in ("MID", "MID & Final")]
    final_set = [r for r in rows if r["included_for"] in ("FINAL", "MID & Final")]

    mid_total_h = sum(r["contact_hours"] for r in mid_set) or 0.0
    final_total_h = sum(r["contact_hours"] for r in final_set) or 0.0

    for r in rows:
        # MID
        if r["included_for"] in ("MID", "MID & Final") and mid_total_h > 0:
            mid_pct = round(100.0 * r["contact_hours"] / mid_total_h, 0)
        else:
            mid_pct = None

        if mid_pct is not None and isinstance(mid_marks, (int, float)):
            mid_w = round(float(mid_marks) * float(mid_pct) / 100.0, 2)
        else:
            mid_w = None

        # FINAL
        if r["included_for"] in ("FINAL", "MID & Final") and final_total_h > 0:
            final_pct = round(100.0 * r["contact_hours"] / final_total_h, 0)
        else:
            final_pct = None

        if final_pct is not None and isinstance(final_marks, (int, float)):
            final_w = round(float(final_marks) * float(final_pct) / 100.0, 2)
        else:
            final_w = None

        r.update({
            "mid_weight_pct": mid_pct,
            "mid_weighted_marks": mid_w,
            "final_weight_pct": final_pct,
            "final_weighted_marks": final_w,
        })

    totals = {
        "mid_total_hours": round(mid_total_h, 2),
        "final_total_hours": round(final_total_h, 2),
        "mid_total_pct": round(sum((r["mid_weight_pct"] or 0) for r in rows), 0),
        "final_total_pct": round(sum((r["final_weight_pct"] or 0) for r in rows), 0),
        "mid_total_weighted": round(sum((r["mid_weighted_marks"] or 0) for r in rows), 2),
        "final_total_weighted": round(sum((r["final_weighted_marks"] or 0) for r in rows), 2),
    }

    return {"rows": rows, "totals": totals}


# ----------------------------
# Excel filling (Sheet 1 only)
# ----------------------------

def _find_first_cell_contains(ws, text: str) -> Optional[Tuple[int, int]]:
    needle = (text or "").strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and needle in v.strip().lower():
                return (cell.row, cell.column)
    return None


def _set_right_of_label(ws, label_contains: str, value: Any) -> None:
    hit = _find_first_cell_contains(ws, label_contains)
    if not hit:
        return
    r, c = hit
    ws.cell(row=r, column=c + 1).value = value


def fill_tos_workbook(template_path: str, output_path: str, model: CourseOutcomesModel) -> str:
    """
    Fills the ToS workbook template with Sheet 1 inputs (COURSE_OUTCOMES).
    Leaves formulas intact (Excel will compute yellow cells when opened).
    """
    wb = load_workbook(template_path)
    ws = wb["COURSE_OUTCOMES"]

    # Header block (labels in your template include colons)
    _set_right_of_label(ws, "Semester:", model.semester)
    _set_right_of_label(ws, "AY:", model.ay)
    _set_right_of_label(ws, "Section:", model.section)

    _set_right_of_label(ws, "Course Code:", model.course_code)
    _set_right_of_label(ws, "Course Title:", model.course_title)
    _set_right_of_label(ws, "Nature of Course:", model.nature_of_course)
    _set_right_of_label(ws, "Level:", model.level)

    _set_right_of_label(ws, "Mid Sem Exam Marks:", model.mid_exam_marks if model.mid_exam_marks is not None else "")
    _set_right_of_label(ws, "Final Exam Marks:", model.final_exam_marks if model.final_exam_marks is not None else "")

    _set_right_of_label(ws, "Template is used for:", model.template_used_for)

    # Find CLO header row (contains "CLO No.")
    header_hit = _find_first_cell_contains(ws, "CLO No.")
    if not header_hit:
        wb.save(output_path)
        return output_path

    header_row, _ = header_hit

    # Determine column letters from the header row
    headers = {}
    for cell in ws[header_row]:
        if isinstance(cell.value, str):
            headers[cell.value.strip()] = cell.column

    col_clo_no = headers.get("CLO No.")
    col_clo_text = headers.get("Course Learning Outcome")
    col_hours = headers.get("Contact hours / Outcome")
    col_included = headers.get("Learning outcome included for")

    start_row = header_row + 1

    # Fill up to 10 rows as the template expects
    for idx in range(10):
        r = start_row + idx
        if idx < len(model.clos):
            clo = model.clos[idx]
            if col_clo_no: ws.cell(row=r, column=col_clo_no).value = clo.clo_no
            if col_clo_text: ws.cell(row=r, column=col_clo_text).value = clo.clo_text
            if col_hours: ws.cell(row=r, column=col_hours).value = float(clo.contact_hours or 0)
            if col_included: ws.cell(row=r, column=col_included).value = _norm_included(clo.included_for)
        else:
            # clear remaining rows
            if col_clo_no: ws.cell(row=r, column=col_clo_no).value = ""
            if col_clo_text: ws.cell(row=r, column=col_clo_text).value = ""
            if col_hours: ws.cell(row=r, column=col_hours).value = ""
            if col_included: ws.cell(row=r, column=col_included).value = ""

    wb.save(output_path)
    return output_path
