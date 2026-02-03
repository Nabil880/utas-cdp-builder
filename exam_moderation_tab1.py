# exam_moderation_tab1.py
from __future__ import annotations

from dataclasses import dataclass
from typing import List, Optional, Dict, Tuple
from io import BytesIO
import json
import re
import math
import datetime as dt

import pandas as pd
import streamlit as st
import openpyxl


# -----------------------------
# Constants (match Excel dropdowns)
# -----------------------------
SEMESTER_OPTIONS = ["", "1", "2", "3"]
SECTION_OPTIONS = ["", "MIE", "EEE", "CAE"]  # "" means unselected
NATURE_OPTIONS = ["Theory", "Theory with Lab", "Practical"]
LEVEL_OPTIONS = ["Bachelor", "Advanced Diploma", "Diploma Second Year", "Diploma First Year"]
TEMPLATE_USED_FOR_OPTIONS = ["", "MID", "FINAL"]
INCLUDED_FOR_OPTIONS = ["", "MID", "MID & Final", "FINAL"]


# -----------------------------
# Models
# -----------------------------
@dataclass
class CLOItem:
    clo_no: int
    clo_text: str = ""
    contact_hours: float = 0.0
    included_for: str = ""  # MID / MID & Final / FINAL

@dataclass
class CourseOutcomesModel:
    semester: str = ""
    ay: str = ""
    section: str = ""  # forced manual selection
    course_code: str = ""
    course_title: str = ""
    nature_of_course: str = ""
    level: str = ""
    mid_sem_marks: Optional[float] = None
    final_exam_marks: Optional[float] = None
    template_used_for: str = ""  # MID / FINAL
    clos: List[CLOItem] = None
    # diagnostics
    unmapped_hours: float = 0.0
    contact_hours_source: str = "theory"  # "theory" or "theory+practical"


# -----------------------------
# Excel-style rounding (ROUND halves away from zero)
# -----------------------------
def excel_round(x: float, ndigits: int) -> float:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return x
    factor = 10 ** ndigits
    if x >= 0:
        return math.floor(x * factor + 0.5) / factor
    return math.ceil(x * factor - 0.5) / factor


# -----------------------------
# JSON parsing helpers
# -----------------------------
def _parse_semester(doc_semester: str) -> str:
    if not doc_semester:
        return ""
    s = doc_semester.strip().lower()
    if "i" in s and "semester" in s:
        # Semester I / II / III
        if "iii" in s:
            return "3"
        if "ii" in s:
            return "2"
        if "i" in s:
            return "1"
    # fallback: try digits
    m = re.search(r"(\d)", s)
    return m.group(1) if m else ""


def _normalize_level(level: str) -> str:
    if not level:
        return ""
    lvl = level.strip().lower()
    mapping = {
        "bachelor": "Bachelor",
        "bsc": "Bachelor",
        "advanced diploma": "Advanced Diploma",
        "adv diploma": "Advanced Diploma",
        "diploma second year": "Diploma Second Year",
        "diploma 2nd year": "Diploma Second Year",
        "diploma first year": "Diploma First Year",
        "diploma 1st year": "Diploma First Year",
    }
    for k, v in mapping.items():
        if k in lvl:
            return v
    # if already matches exactly, keep it
    if level in LEVEL_OPTIONS:
        return level
    return level  # last resort (still editable in UI)


def _derive_nature(hours_theory: float, hours_practical: float) -> str:
    ht = float(hours_theory or 0)
    hp = float(hours_practical or 0)
    if ht > 0 and hp > 0:
        return "Theory with Lab"
    if hp > 0 and ht == 0:
        return "Practical"
    return "Theory"


def _clo_num(label: str) -> Optional[int]:
    if not label:
        return None
    m = re.search(r"(\d+)", str(label))
    return int(m.group(1)) if m else None


def _compute_contact_hours_from_distribution(
    cdp: dict,
    source: str = "theory",  # "theory" or "theory+practical"
) -> Tuple[Dict[int, float], float]:
    """
    Default rule (matches Excel-friendly integers):
      - If a topic row is mapped to multiple CLOs, each CLO gets the full row hours.
      - For exam ToS, we default to theory rows only.
    """
    totals: Dict[int, float] = {}
    unmapped = 0.0

    keys = ["theory_df"] if source == "theory" else ["theory_df", "practical_df"]

    for k in keys:
        for row in cdp.get(k, []) or []:
            hrs = float(row.get("hours") or 0)
            clos = row.get("clos") or []
            clo_nums = [n for n in (_clo_num(c) for c in clos) if n is not None]

            if not clo_nums:
                unmapped += hrs
                continue

            for n in clo_nums:
                totals[n] = totals.get(n, 0.0) + hrs

    return totals, unmapped


def extract_course_outcomes_from_cdp(cdp: dict, contact_hours_source: str = "theory") -> CourseOutcomesModel:
    course = cdp.get("course", {}) or {}
    doc = cdp.get("doc", {}) or {}

    model = CourseOutcomesModel()
    model.semester = _parse_semester(doc.get("semester", ""))
    model.ay = str(doc.get("academic_year", "") or "")
    model.section = ""  # forced manual selection
    model.course_code = str(course.get("code", "") or "")
    model.course_title = str(course.get("title", "") or "")
    model.nature_of_course = _derive_nature(course.get("hours_theory"), course.get("hours_practical"))
    model.level = _normalize_level(course.get("level", ""))
    model.mid_sem_marks = None
    model.final_exam_marks = None
    model.template_used_for = ""
    model.contact_hours_source = contact_hours_source

    # CLO list
    clos_df = cdp.get("clos_df", []) or []
    clos_items: List[CLOItem] = []

    # sort by clo_no if present, else keep order
    def clo_sort_key(d: dict, idx: int) -> int:
        cn = d.get("clo_no")
        try:
            return int(cn)
        except Exception:
            return idx + 1

    clos_df_sorted = sorted(list(enumerate(clos_df)), key=lambda t: clo_sort_key(t[1], t[0]))
    for i, clo_row in clos_df_sorted[:10]:
        cn = clo_row.get("clo_no")
        try:
            clo_no = int(cn)
        except Exception:
            clo_no = len(clos_items) + 1
        clo_text = str(clo_row.get("learning_outcomes", "") or "").strip()
        clos_items.append(CLOItem(clo_no=clo_no, clo_text=clo_text, contact_hours=0.0, included_for=""))

    # contact hours by CLO number
    hours_map, unmapped = _compute_contact_hours_from_distribution(cdp, source=contact_hours_source)
    model.unmapped_hours = float(unmapped)

    # attach hours to the CLO list
    for item in clos_items:
        item.contact_hours = float(hours_map.get(item.clo_no, 0.0))

    model.clos = clos_items
    return model


# -----------------------------
# Computation spec (for preview)
# -----------------------------
def compute_tos_numbers(model: CourseOutcomesModel) -> pd.DataFrame:
    rows = []
    # Use only CLO rows that have either text or hours (keeps preview clean)
    clos = [c for c in (model.clos or []) if (str(c.clo_text).strip() or float(c.contact_hours or 0) > 0)]

    mid_marks_total = model.mid_sem_marks if model.mid_sem_marks is not None else None
    final_marks_total = model.final_exam_marks if model.final_exam_marks is not None else None

    mid_total = sum(
        float(c.contact_hours or 0)
        for c in clos
        if c.included_for in ("MID", "MID & Final")
    )
    final_total = sum(float(c.contact_hours or 0) for c in clos)

    # Compute per CLO
    mid_sum_marks = 0.0
    final_sum_marks = 0.0
    mid_sum_pct = 0.0
    final_sum_pct = 0.0

    for c in clos:
        contact = float(c.contact_hours or 0)

        # Mid
        if mid_total > 0 and c.included_for in ("MID", "MID & Final") and contact > 0 and mid_marks_total is not None:
            mid_w = contact / mid_total
            mid_marks = mid_w * mid_marks_total
            mid_marks_2dp = excel_round(mid_marks, 2)
            weighted_mid = excel_round(mid_marks_2dp, 0)
            mid_pct_disp = f"{excel_round(mid_w * 100, 0):.0f}%"
            mid_sum_marks += float(weighted_mid)
            mid_sum_pct += mid_w
        else:
            mid_w = None
            weighted_mid = None
            mid_pct_disp = ""

        # Final
        # Template subtracts mid_marks (unrounded) from final allocation.
        mid_marks_for_sub = (mid_w * mid_marks_total) if (mid_w is not None and mid_marks_total is not None) else 0.0

        if final_total > 0 and contact > 0 and final_marks_total is not None:
            final_w = contact / final_total
            final_marks = final_w * final_marks_total - mid_marks_for_sub
            final_marks_2dp = excel_round(final_marks, 2)
            weighted_final = excel_round(final_marks_2dp, 0)
            final_pct_disp = f"{excel_round(final_w * 100, 0):.0f}%"
            final_sum_marks += float(weighted_final)
            final_sum_pct += final_w
        else:
            final_w = None
            weighted_final = None
            final_pct_disp = ""

        rows.append({
            "CLO No.": c.clo_no,
            "Course Learning Outcome": c.clo_text,
            "Contact hours / Outcome": contact,
            "Included For": c.included_for,
            "% Mid": mid_pct_disp,
            "Weighted Marks Mid": (f"{float(weighted_mid):.2f}" if weighted_mid is not None else ""),
            "% Final": final_pct_disp,
            "Weighted Marks Final": (f"{float(weighted_final):.1f}" if weighted_final is not None else ""),
        })

    df = pd.DataFrame(rows)

    # totals row (preview)
    total_row = {
        "CLO No.": "",
        "Course Learning Outcome": "TOTAL",
        "Contact hours / Outcome": sum(float(c.contact_hours or 0) for c in clos),
        "Included For": "",
        "% Mid": f"{excel_round(mid_sum_pct * 100, 0):.0f}%" if mid_marks_total is not None and mid_total > 0 else "",
        "Weighted Marks Mid": f"{mid_sum_marks:.2f}" if mid_marks_total is not None and mid_total > 0 else "",
        "% Final": f"{excel_round(final_sum_pct * 100, 0):.0f}%" if final_marks_total is not None and final_total > 0 else "",
        "Weighted Marks Final": f"{final_sum_marks:.1f}" if final_marks_total is not None and final_total > 0 else "",
    }
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    return df


# -----------------------------
# Fill workbook (only blue fields; keep template formulas)
# -----------------------------
def fill_tos_workbook(template_bytes: bytes, model: CourseOutcomesModel) -> bytes:
    bio = BytesIO(template_bytes)
    wb = openpyxl.load_workbook(bio, data_only=False)
    ws = wb["COURSE_OUTCOMES"]

    # Header cells (exact positions from your template)
    ws["E5"].value = model.semester or ""
    ws["G5"].value = model.ay or ""
    ws["I5"].value = model.section or ""  # forced manual select
    ws["E7"].value = model.course_code or ""
    ws["E8"].value = model.course_title or ""
    ws["E9"].value = model.nature_of_course or ""
    ws["E10"].value = model.level or ""
    ws["I10"].value = "" if model.mid_sem_marks is None else float(model.mid_sem_marks)
    ws["I12"].value = "" if model.final_exam_marks is None else float(model.final_exam_marks)
    ws["D12"].value = model.template_used_for or ""

    # CLO rows (16..25)
    start_row = 16
    max_rows = 10
    clos = model.clos or []

    for i in range(max_rows):
        r = start_row + i
        if i < len(clos):
            c = clos[i]
            ws[f"C{r}"].value = c.clo_text or ""
            # Contact hours field in template is formatted "0" (integer display)
            ws[f"D{r}"].value = float(c.contact_hours or 0.0)
            ws[f"I{r}"].value = c.included_for or ""
        else:
            ws[f"C{r}"].value = ""
            ws[f"D{r}"].value = ""
            ws[f"I{r}"].value = ""

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# -----------------------------
# Streamlit UI
# -----------------------------
@st.cache_data
def load_template_bytes(path: str) -> bytes:
    with open(path, "rb") as f:
        return f.read()


def exam_moderation_tab1(TOS_TEMPLATE_PATH: str):
    st.subheader("Exam Moderation — COURSE_OUTCOMES (Tab 1)")

    # Upload approved JSON (signature verification comes next milestone)
    uploaded = st.file_uploader("Upload Approved CDP JSON", type=["json"], key="tos_json_upload")

    # Contact-hour source choice (default: theory-only for exam ToS)
    contact_source = st.radio(
        "Contact hours source for CLO weighting",
        options=["theory", "theory+practical"],
        index=0,
        help="Default is theory-only because ToS is for mid/final exams. Switch to include practical hours if needed.",
        horizontal=True
    )

    if uploaded:
        cdp = json.load(uploaded)
        model = extract_course_outcomes_from_cdp(cdp, contact_hours_source=contact_source)

        # Keep in session so edits persist
        if "tos_model" not in st.session_state:
            st.session_state.tos_model = model
        else:
            # If user re-uploads a new JSON, reset model
            st.session_state.tos_model = model

    if "tos_model" not in st.session_state:
        st.info("Upload an approved CDP JSON to auto-fill the COURSE_OUTCOMES inputs.")
        return

    model: CourseOutcomesModel = st.session_state.tos_model

    # --- Header editor ---
    st.markdown("### Header fields (match Excel dropdowns)")
    c1, c2, c3 = st.columns(3)
    with c1:
        model.semester = st.selectbox("Semester", SEMESTER_OPTIONS, index=SEMESTER_OPTIONS.index(model.semester or ""))
        model.ay = st.text_input("AY", value=model.ay or "")
        model.course_code = st.text_input("Course Code", value=model.course_code or "")
    with c2:
        # SECTION MUST BE MANUAL (your preference)
        model.section = st.selectbox("Section (required)", SECTION_OPTIONS, index=0)
        model.course_title = st.text_input("Course Title", value=model.course_title or "")
        model.level = st.selectbox("Level", LEVEL_OPTIONS, index=LEVEL_OPTIONS.index(model.level) if model.level in LEVEL_OPTIONS else 0)
    with c3:
        model.nature_of_course = st.selectbox(
            "Nature of Course",
            NATURE_OPTIONS,
            index=NATURE_OPTIONS.index(model.nature_of_course) if model.nature_of_course in NATURE_OPTIONS else 0,
        )
        model.template_used_for = st.selectbox("Template is used for", TEMPLATE_USED_FOR_OPTIONS, index=TEMPLATE_USED_FOR_OPTIONS.index(model.template_used_for or ""))
        model.mid_sem_marks = st.number_input("Mid Sem Exam Marks", min_value=0.0, step=1.0, value=float(model.mid_sem_marks or 0.0))
        model.final_exam_marks = st.number_input("Final Exam Marks", min_value=0.0, step=1.0, value=float(model.final_exam_marks or 0.0))

    # --- CLO grid editor ---
    st.markdown("### CLO inputs (blue fields)")
    clo_rows = []
    for c in (model.clos or []):
        clo_rows.append({
            "CLO No.": c.clo_no,
            "Course Learning Outcome": c.clo_text,
            "Contact hours / Outcome": float(c.contact_hours or 0.0),
            "Learning outcome included for": c.included_for,
        })
    df_edit = pd.DataFrame(clo_rows)

    edited = st.data_editor(
        df_edit,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config={
            "CLO No.": st.column_config.NumberColumn("CLO No.", disabled=True),
            "Course Learning Outcome": st.column_config.TextColumn("Course Learning Outcome", width="large"),
            "Contact hours / Outcome": st.column_config.NumberColumn("Contact hours / Outcome", min_value=0.0, step=1.0),
            "Learning outcome included for": st.column_config.SelectboxColumn(
                "Learning outcome included for",
                options=INCLUDED_FOR_OPTIONS,
                required=False,
            ),
        },
        disabled=["CLO No."],
        key="tos_clo_editor",
    )

    # Write edits back to model
    new_clos: List[CLOItem] = []
    for _, row in edited.iterrows():
        new_clos.append(CLOItem(
            clo_no=int(row["CLO No."]),
            clo_text=str(row["Course Learning Outcome"] or ""),
            contact_hours=float(row["Contact hours / Outcome"] or 0.0),
            included_for=str(row["Learning outcome included for"] or ""),
        ))
    model.clos = new_clos

    # Diagnostics
    if model.unmapped_hours > 0:
        st.warning(f"Weekly distribution contains ~{model.unmapped_hours:.1f} hour(s) with no CLO mapping. They are not counted in contact hours.")

    # --- Computed preview ---
    st.markdown("### Computed preview (yellow fields logic)")
    preview_df = compute_tos_numbers(model)
    st.dataframe(preview_df, use_container_width=True)

    # --- Export ---
    st.markdown("### Export")
    errors = []
    if not model.section:
        errors.append("Section is required (must be selected manually).")
    if not model.semester:
        errors.append("Semester is required.")
    if not model.course_code.strip():
        errors.append("Course Code is required.")
    if not model.course_title.strip():
        errors.append("Course Title is required.")
    if model.template_used_for not in ("MID", "FINAL"):
        errors.append("Template is used for must be MID or FINAL.")

    if errors:
        st.error("Fix these before export:\n- " + "\n- ".join(errors))

    template_bytes = load_template_bytes(TOS_TEMPLATE_PATH)

    if st.button("Generate ToS Workbook (Sheet 1 filled)", type="primary", disabled=bool(errors)):
        out_bytes = fill_tos_workbook(template_bytes, model)
        filename = f"{model.course_code}_ToS_COURSE_OUTCOMES_{dt.date.today().isoformat()}.xlsx"
        st.download_button(
            "Download Excel",
            data=out_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.success("Workbook generated from the official template (formulas preserved).")
